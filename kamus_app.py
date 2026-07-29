# kamus_app.py
# Streamlit app: Kamus Arab-Indonesia PDF → Excel via Google Gemini AI
# Mendukung: PDF digital (teks) & PDF scan (gambar/foto)
# Cara jalankan: streamlit run kamus_app.py

import streamlit as st
import google.generativeai as genai
import fitz          # PyMuPDF
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageEnhance, ImageFilter
import json, io, time, re
from datetime import datetime

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG
# ═══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="📖 Kamus Arab-Indonesia Parser",
    page_icon="📖",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ═══════════════════════════════════════════════════════════════════════════════
# CSS
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+Arabic:wght@400;700&family=Merriweather:ital,wght@0,400;0,700;1,400&family=Inter:wght@400;500;600&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.stApp { background: #F9F5EE; }

.app-header {
    background: linear-gradient(135deg, #1C2B3A 0%, #2E4057 100%);
    color: white; padding: 28px 36px 22px; border-radius: 12px;
    margin-bottom: 24px; border-bottom: 4px solid #C9A84C;
    box-shadow: 0 4px 20px rgba(0,0,0,0.15);
}
.app-title {
    font-family: 'Merriweather', serif; font-size: 1.9em; font-weight: 700;
    margin: 0 0 4px 0; letter-spacing: -0.3px;
}
.app-sub { color: #C9A84C; font-size: 0.9em; letter-spacing: 0.8px;
           text-transform: uppercase; font-weight: 500; }

/* PDF type badge */
.pdf-badge {
    display: inline-flex; align-items: center; gap: 6px;
    padding: 5px 14px; border-radius: 20px; font-size: 0.85em;
    font-weight: 700; letter-spacing: 0.3px; margin: 6px 0;
}
.pdf-digital { background: #DBEAFE; color: #1E40AF; border: 1.5px solid #93C5FD; }
.pdf-scan    { background: #FCE7F3; color: #9D174D; border: 1.5px solid #F9A8D4; }
.pdf-mixed   { background: #FEF3C7; color: #92400E; border: 1.5px solid #FCD34D; }

/* Live feed */
.feed-wrap {
    max-height: 520px; overflow-y: auto; padding: 6px 2px;
    scrollbar-width: thin; scrollbar-color: #C9A84C #F0EBE1;
}
.entry-card {
    background: #FFFFFF; border: 1px solid #E8E0D0;
    border-left: 5px solid #2E4057; padding: 10px 14px 8px;
    margin: 6px 0; border-radius: 6px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
}
.entry-card.sublema {
    border-left-color: #C9A84C; margin-left: 18px; background: #FDFAF4;
}
.badge {
    display: inline-block; padding: 2px 9px; border-radius: 3px;
    font-size: 0.68em; font-weight: 700; letter-spacing: 0.8px;
    margin-right: 8px; text-transform: uppercase; vertical-align: middle;
}
.badge-lema    { background: #2E4057; color: #FFF; }
.badge-sublema { background: #C9A84C; color: #1C2B3A; }

.arab-text {
    font-family: 'Noto Sans Arabic', 'Traditional Arabic', serif;
    font-size: 1.35em; direction: rtl; color: #1C2B3A;
    font-weight: 700; vertical-align: middle; line-height: 1.8;
}
.card-detail {
    margin-top: 3px; font-size: 0.82em; color: #555;
    display: flex; flex-wrap: wrap; gap: 8px; align-items: center;
}
.pill { background:#EEF2F7; color:#2E4057; padding:1px 8px; border-radius:10px; font-size:0.92em; }
.pill.trans { background:#FFF8E7; color:#7A5C00; font-style:italic; }
.pill.eku   { background:#E8F5E9; color:#1B5E20; font-weight:600; }
.pill.arti  { background:#EDE7F6; color:#4527A0; }
.pill.alt   { background:#FCE4EC; color:#880E4F; }

/* Stats */
.stat-strip { display:flex; gap:12px; margin:16px 0; }
.stat-box {
    flex:1; background:#2E4057; color:white; padding:12px 10px;
    border-radius:8px; text-align:center; border-bottom:3px solid #C9A84C;
}
.stat-num   { font-size:1.8em; font-weight:700; line-height:1.1; }
.stat-label { font-size:0.72em; opacity:0.75; text-transform:uppercase; letter-spacing:0.5px; }

/* Status */
.status-pill {
    display:inline-flex; align-items:center; gap:6px;
    padding:5px 14px; border-radius:20px; font-size:0.85em; font-weight:600;
}
.status-idle       { background:#E8ECF0; color:#2E4057; }
.status-processing { background:#FFF8DC; color:#7A5C00; }
.status-done       { background:#D6F0D8; color:#1B5E20; }

.section-title {
    font-family:'Merriweather',serif; font-size:1.05em; color:#2E4057;
    font-weight:700; border-bottom:2px solid #C9A84C; padding-bottom:4px; margin:16px 0 10px;
}

div[data-testid="stMetricValue"] { font-size:1.6em; color:#1C2B3A; }
.stProgress > div > div { background:#C9A84C !important; }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ═══════════════════════════════════════════════════════════════════════════════
for key, default in [
    ('entries',   []),
    ('feed',      []),
    ('done',      False),
    ('pdf_bytes', None),
    ('pdf_name',  ''),
    ('n_pages',   0),
    ('pdf_type',  None),   # 'digital' | 'scan' | 'mixed'
    ('scan_ratio', 0.0),
]:
    if key not in st.session_state:
        st.session_state[key] = default

# ═══════════════════════════════════════════════════════════════════════════════
# HELPER: DETEKSI TIPE PDF
# ═══════════════════════════════════════════════════════════════════════════════

def detect_pdf_type(pdf_bytes: bytes) -> tuple:
    """
    Deteksi apakah PDF berisi teks digital, hasil scan, atau campuran.
    Return: (tipe: str, scan_ratio: float, detail: str)
    """
    doc = fitz.open(stream=pdf_bytes, filetype='pdf')
    total = len(doc)
    sample_n = min(8, total)   # sample maks 8 halaman
    
    text_pages = 0
    scan_pages = 0
    
    for i in range(sample_n):
        page = doc[i]
        text = page.get_text().strip()
        
        # Hitung jumlah karakter teks yang bisa diekstrak
        # Halaman digital biasanya punya > 100 karakter teks
        if len(text) >= 80:
            text_pages += 1
        else:
            # Cek apakah ada image di halaman ini
            imgs = page.get_images(full=False)
            if imgs:
                scan_pages += 1
            else:
                # Halaman kosong / dekoratif
                text_pages += 1
    
    doc.close()
    
    scan_ratio = scan_pages / sample_n if sample_n > 0 else 0
    
    if scan_ratio >= 0.8:
        return 'scan', scan_ratio, f"Terdeteksi PDF scan/foto ({int(scan_ratio*100)}% halaman)"
    elif scan_ratio >= 0.2:
        return 'mixed', scan_ratio, f"PDF campuran: {scan_pages} scan + {text_pages} digital"
    else:
        return 'digital', scan_ratio, f"PDF digital dengan teks tersemat"


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER: PREPROCESSING GAMBAR UNTUK SCAN
# ═══════════════════════════════════════════════════════════════════════════════

def preprocess_for_scan(img: Image.Image, mode: str = 'balanced') -> Image.Image:
    """
    Tingkatkan kualitas gambar scan agar Gemini bisa membaca lebih akurat.
    
    mode:
      'light'    — peningkatan minimal, untuk scan yang sudah cukup jelas
      'balanced' — standar, untuk scan biasa
      'strong'   — agresif, untuk scan buram / kontras rendah
    """
    if img.mode not in ('RGB', 'L'):
        img = img.convert('RGB')
    
    if mode == 'light':
        contrast_f  = 1.2
        sharpness_f = 1.2
        bright_f    = 1.0
    elif mode == 'strong':
        contrast_f  = 1.6
        sharpness_f = 1.8
        bright_f    = 1.1
    else:  # balanced
        contrast_f  = 1.35
        sharpness_f = 1.4
        bright_f    = 1.03

    img = ImageEnhance.Contrast(img).enhance(contrast_f)
    img = ImageEnhance.Sharpness(img).enhance(sharpness_f)
    img = ImageEnhance.Brightness(img).enhance(bright_f)

    # Filter unsharp mask ringan untuk memperjelas tepi huruf
    img = img.filter(ImageFilter.UnsharpMask(radius=1, percent=80, threshold=3))

    return img


def auto_rotate(img: Image.Image) -> Image.Image:
    """
    Koreksi rotasi gambar berdasarkan EXIF jika ada.
    Berguna untuk scan dari kamera/HP yang memiliki metadata orientasi.
    """
    try:
        from PIL import ExifTags
        exif = img._getexif()
        if exif:
            for tag, val in exif.items():
                if ExifTags.TAGS.get(tag) == 'Orientation':
                    if val == 3:
                        img = img.rotate(180, expand=True)
                    elif val == 6:
                        img = img.rotate(270, expand=True)
                    elif val == 8:
                        img = img.rotate(90, expand=True)
    except Exception:
        pass
    return img


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER: KONVERSI PDF → GAMBAR
# ═══════════════════════════════════════════════════════════════════════════════

def pdf_pages_to_images(
    pdf_bytes: bytes,
    start: int,
    end: int,
    dpi: int,
    is_scan: bool = False,
    preprocess_mode: str = 'balanced',
) -> list:
    """
    Konversi halaman PDF ke list PIL Image.
    Untuk scan, terapkan preprocessing peningkatan kualitas.
    """
    doc = fitz.open(stream=pdf_bytes, filetype='pdf')
    imgs = []

    for i in range(start, min(end, len(doc))):
        page = doc[i]

        # Render halaman ke gambar piksel
        pix = page.get_pixmap(dpi=dpi)
        img = Image.open(io.BytesIO(pix.tobytes("png")))

        # Auto-rotate (berguna untuk scan dari kamera)
        img = auto_rotate(img)

        # Preprocessing khusus scan
        if is_scan and preprocess_mode != 'none':
            img = preprocess_for_scan(img, mode=preprocess_mode)

        imgs.append(img)

    doc.close()
    return imgs


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER: PROMPT GEMINI
# ═══════════════════════════════════════════════════════════════════════════════

def build_prompt(n_pages: int, is_scan: bool = False) -> str:

    ocr_block = ""
    if is_scan:
        ocr_block = """
═══ MODE OCR — PDF INI ADALAH HASIL SCAN ═══
  • Halaman berikut adalah gambar/foto fisik buku, bukan PDF digital.
  • Lakukan OCR (baca teks langsung dari gambar) seakurat mungkin.
  • Aksara Arab dengan harakat: tulis lengkap jika terbaca, tanpa harakat jika buram.
  • Jika ada bagian teks tidak terbaca (tertutup, robek, buram): kosongkan field itu.
  • Abaikan artefak scan: bintik, bayangan jari, garis lipatan, noise latar belakang.
  • Halaman miring ringan: tetap baca, jangan lewati.
"""

    return f"""Kamu adalah parser ahli kamus Indonesia-Arab yang sangat teliti dan akurat.

Berikut adalah {n_pages} halaman dari kamus Indonesia-Arab.
Baca SEMUA kolom dari kiri ke kanan, atas ke bawah. Jangan lewati satu entri pun.
{ocr_block}
═══ FORMAT ENTRI DI KAMUS ═══
  kata_indonesia (keterangan) [aksara_arab] transliterasi_latin

═══ CARA MEMBEDAKAN LEMA vs SUBLEMA ═══
  • LEMA    = entri utama — posisi paling kiri di kolom, biasanya tebal/bold
  • SUBLEMA = entri turunan — menjorok ke kanan, atau diawali ~, atau diawali ::

═══ ATURAN KOLOM (WAJIB DIIKUTI KETAT) ═══
  1. LEMA          → HANYA aksara Arab. Kosong jika baris ini adalah SUBLEMA.
  2. SUBLEMA       → HANYA aksara Arab. Kosong jika baris ini adalah LEMA.
  3. TRANSLITERASI → HANYA huruf Latin (romanisasi Arab). TANPA aksara Arab.
  4. EKUIVALEN     → kata Indonesia MAKSIMAL 2 KATA sebelum [aksara_arab].
                     TIDAK BOLEH berisi aksara Arab. TIDAK BOLEH berisi transliterasi Latin.
  5. ALTERNATIF    → arti Indonesia lain jika ada pilihan ("atau" / "/"). Kosong jika tidak ada.
  6. ARTI          → keterangan lebih dari 2 kata, atau penjelasan dalam tanda kurung (...).
                     TIDAK BOLEH berisi aksara Arab.
  7. SINONIM       → sinonim bahasa Arab atau Indonesia jika ada. Kosong jika tidak ada.

═══ CONTOH PARSING ═══
  "abad [قَرْنٌ] qarnun ج"
  → {{"LEMA":"قَرْنٌ","SUBLEMA":"","TRANSLITERASI":"qarnun","EKUIVALEN":"abad","ALTERNATIF":"","ARTI":"","SINONIM":""}}

  "  berabad-abad [لِعِدَّةِ قُرُونٍ] li 'iddati quruunin"
  → {{"LEMA":"","SUBLEMA":"لِعِدَّةِ قُرُونٍ","TRANSLITERASI":"li 'iddati quruunin","EKUIVALEN":"berabad-abad","ALTERNATIF":"","ARTI":"","SINONIM":""}}

  "abai (me-kan) [أَهْمَلَ] ahmala- [تَرَكَ] taraka ~"
  → 2 baris:
    {{"LEMA":"أَهْمَلَ","SUBLEMA":"","TRANSLITERASI":"ahmala-","EKUIVALEN":"abai","ALTERNATIF":"","ARTI":"(me-kan)","SINONIM":""}}
    {{"LEMA":"","SUBLEMA":"تَرَكَ","TRANSLITERASI":"taraka","EKUIVALEN":"abai","ALTERNATIF":"","ARTI":"","SINONIM":""}}

  "abdi [عَبْدٌ] 'abdun- ~ (pe-an) [خِدْمَةٌ] khidmatun"
  → 2 baris:
    {{"LEMA":"عَبْدٌ","SUBLEMA":"","TRANSLITERASI":"'abdun-","EKUIVALEN":"abdi","ALTERNATIF":"","ARTI":"","SINONIM":""}}
    {{"LEMA":"","SUBLEMA":"خِدْمَةٌ","TRANSLITERASI":"khidmatun","EKUIVALEN":"","ALTERNATIF":"","ARTI":"(pe-an)","SINONIM":""}}

═══ ABAIKAN ═══
  Nomor halaman, header huruf (A B C ...), garis dekoratif, daftar isi, watermark.

RETURN: HANYA array JSON, tanpa penjelasan, tanpa markdown, tanpa komentar apapun.
[{{"LEMA":"","SUBLEMA":"","TRANSLITERASI":"","EKUIVALEN":"","ALTERNATIF":"","ARTI":"","SINONIM":""}}, ...]"""


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER: VALIDASI ENTRI
# ═══════════════════════════════════════════════════════════════════════════════

ARAB_RE = re.compile(r'[\u0600-\u06FF\u0750-\u077F\uFB50-\uFDFF\uFE70-\uFEFF]')

def clean_arab(v):
    v = str(v or '').strip()
    v = re.sub(r'[A-Za-z0-9\.\,\;\:\!\?\-\/\\]+', '', v).strip()
    return v

def clean_latin(v):
    v = str(v or '').strip()
    return ARAB_RE.sub('', v).strip()

def clean_indo(v):
    v = str(v or '').strip()
    return ARAB_RE.sub('', v).strip()

def validate_entry(e: dict) -> dict:
    row = {
        'LEMA':          clean_arab(e.get('LEMA',  '')),
        'SUBLEMA':       clean_arab(e.get('SUBLEMA', '')),
        'TRANSLITERASI': clean_latin(e.get('TRANSLITERASI', '')),
        'EKUIVALEN':     clean_indo(e.get('EKUIVALEN', '')),
        'ALTERNATIF':    clean_indo(e.get('ALTERNATIF', '')),
        'ARTI':          clean_indo(e.get('ARTI', '')),
        'SINONIM':       str(e.get('SINONIM', '') or '').strip(),
    }
    if row['LEMA'] and row['SUBLEMA']:
        row['SUBLEMA'] = ''
    return row

def is_valid_entry(row: dict) -> bool:
    arab = row['LEMA'] or row['SUBLEMA']
    if not arab or len(arab) < 2:
        return False
    eku = row['EKUIVALEN']
    if len(eku) == 1 and eku.isalpha():
        return False
    return True


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER: PANGGIL GEMINI
# ═══════════════════════════════════════════════════════════════════════════════

def call_gemini(model, images: list, is_scan: bool = False, retries: int = 2) -> list:
    prompt = build_prompt(len(images), is_scan=is_scan)

    for attempt in range(retries):
        try:
            resp = model.generate_content(
                [prompt] + images,
                generation_config=genai.GenerationConfig(
                    temperature=0.05,
                    max_output_tokens=8192,
                )
            )
            raw = resp.text.strip()
            raw = re.sub(r'^```(?:json)?\s*', '', raw)
            raw = re.sub(r'\s*```$', '', raw).strip()

            if not raw.startswith('['):
                m = re.search(r'\[.*\]', raw, re.DOTALL)
                raw = m.group(0) if m else '[]'

            data = json.loads(raw)
            results = []
            for e in data:
                if not isinstance(e, dict):
                    continue
                row = validate_entry(e)
                if is_valid_entry(row):
                    results.append(row)
            return results

        except json.JSONDecodeError as ex:
            if attempt < retries - 1:
                time.sleep(3)
                continue
            st.warning(f"⚠️ JSON error: {str(ex)[:80]}")
            return []
        except Exception as ex:
            msg = str(ex)
            if attempt < retries - 1:
                time.sleep(5)
                continue
            st.error(f"❌ Gemini error: {msg[:200]}")
            return []
    return []


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER: UI CARD & EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def entry_card_html(e: dict) -> str:
    is_lema = bool(e.get('LEMA'))
    arab    = e.get('LEMA') or e.get('SUBLEMA', '')
    trans   = e.get('TRANSLITERASI', '')
    eku     = e.get('EKUIVALEN', '')
    alt     = e.get('ALTERNATIF', '')
    arti    = e.get('ARTI', '')
    cls     = "entry-card" if is_lema else "entry-card sublema"
    badge   = f'<span class="badge badge-{"lema" if is_lema else "sublema"}">{"LEMA" if is_lema else "SUBLEMA"}</span>'
    pills   = ""
    if eku:   pills += f'<span class="pill eku">📝 {eku}</span>'
    if trans: pills += f'<span class="pill trans">🔤 {trans}</span>'
    if alt:   pills += f'<span class="pill alt">≈ {alt}</span>'
    if arti:  pills += f'<span class="pill arti">ℹ {arti}</span>'
    return f'<div class="{cls}">{badge}<span class="arab-text">{arab}</span><div class="card-detail">{pills}</div></div>'


def make_excel(entries: list) -> bytes:
    COLS = ['LEMA','SUBLEMA','TRANSLITERASI','EKUIVALEN','ALTERNATIF','ARTI','SINONIM']
    df = pd.DataFrame(entries, columns=COLS) if entries else pd.DataFrame(columns=COLS)
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine='openpyxl')
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active

    hfill = PatternFill('solid', fgColor='1C2B3A')
    hfont = Font(name='Calibri', bold=True, color='C9A84C', size=11)
    for cell in ws[1]:
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col, w in zip('ABCDEFG', [22, 22, 25, 20, 20, 42, 18]):
        ws.column_dimensions[col].width = w

    f_arab = Font(name='Noto Sans Arabic', size=13)
    f_ital = Font(name='Calibri', size=11, italic=True)
    f_body = Font(name='Calibri', size=11)
    z1 = PatternFill('solid', fgColor='F9F5EE')
    z2 = PatternFill('solid', fgColor='FFFFFF')

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ridx = row[0].row
        for cell in row:
            cl = cell.column_letter
            cell.fill = z1 if ridx % 2 == 0 else z2
            if cl in ('A','B'):
                if cell.value: cell.font = f_arab
                cell.alignment = Alignment(horizontal='right', vertical='center', reading_order=2)
            elif cl == 'C':
                cell.font = f_ital
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.font = f_body
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(cl=='F'))

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def stats_html_render(entries, page_info=""):
    n_l = sum(1 for e in entries if e.get('LEMA'))
    n_s = sum(1 for e in entries if e.get('SUBLEMA'))
    return f"""<div class="stat-strip">
  <div class="stat-box"><div class="stat-num">{len(entries)}</div><div class="stat-label">Total Entri</div></div>
  <div class="stat-box"><div class="stat-num">{n_l}</div><div class="stat-label">LEMA</div></div>
  <div class="stat-box"><div class="stat-num">{n_s}</div><div class="stat-label">SUBLEMA</div></div>
  <div class="stat-box"><div class="stat-num">{page_info or st.session_state.n_pages}</div><div class="stat-label">Halaman</div></div>
</div>"""


# ═══════════════════════════════════════════════════════════════════════════════
# UI — HEADER
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="app-header">
  <div class="app-title">📖 Kamus Arab-Indonesia Parser</div>
  <div class="app-sub">PDF Digital &amp; Scan → Excel · Powered by Google Gemini AI</div>
</div>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### ⚙️ Konfigurasi")

    api_key = st.text_input(
        "🔑 Gemini API Key", type="password", placeholder="AIza...",
        help="Buat di https://aistudio.google.com/app/apikey",
    )

    model_name = st.selectbox(
        "🤖 Model Gemini",
        ["gemini-3.5-flash", "gemini-3.1-flash-lite", "gemini-3.1-pro"],
        help="3.5-flash = tercepat & terbaru | 3.1-pro = paling akurat untuk scan buram",
    )

    st.markdown("---")

    # ── Upload PDF ──────────────────────────────────────────────────────────
    uploaded = st.file_uploader(
        "📄 Upload PDF Kamus",
        type=["pdf"],
        help="Mendukung PDF digital (teks) maupun PDF scan (gambar/foto)",
    )

    if uploaded:
        if st.session_state.pdf_name != uploaded.name:
            with st.spinner("🔍 Menganalisis tipe PDF..."):
                st.session_state.pdf_bytes = uploaded.read()
                st.session_state.pdf_name  = uploaded.name

                doc_tmp = fitz.open(stream=st.session_state.pdf_bytes, filetype='pdf')
                st.session_state.n_pages = len(doc_tmp)
                doc_tmp.close()

                pdf_type, scan_ratio, detail = detect_pdf_type(st.session_state.pdf_bytes)
                st.session_state.pdf_type   = pdf_type
                st.session_state.scan_ratio = scan_ratio

                # Reset hasil lama
                st.session_state.entries = []
                st.session_state.feed    = []
                st.session_state.done    = False

        # Tampilkan info PDF
        n_pages  = st.session_state.n_pages
        pdf_type = st.session_state.pdf_type

        if pdf_type == 'scan':
            st.markdown(
                f'<div class="pdf-badge pdf-scan">📷 PDF Scan / Foto</div>',
                unsafe_allow_html=True,
            )
            st.caption("Mode OCR aktif otomatis")
        elif pdf_type == 'mixed':
            st.markdown(
                f'<div class="pdf-badge pdf-mixed">📋 PDF Campuran</div>',
                unsafe_allow_html=True,
            )
            st.caption("Sebagian scan, sebagian digital")
        else:
            st.markdown(
                f'<div class="pdf-badge pdf-digital">📄 PDF Digital</div>',
                unsafe_allow_html=True,
            )
            st.caption("Teks tersemat — akurasi optimal")

        st.success(f"✅ **{n_pages} halaman** dimuat")
        st.caption(f"Ukuran: {round(len(st.session_state.pdf_bytes)/1024/1024,1)} MB")

    st.markdown("---")

    # ── Setting batch ───────────────────────────────────────────────────────
    batch_size = st.select_slider(
        "📦 Halaman per batch",
        options=[3, 5, 7, 10], value=5,
        help="3-5 = lebih akurat | 7-10 = lebih cepat",
    )

    # DPI: auto-suggest lebih tinggi untuk scan
    default_dpi = 200 if st.session_state.pdf_type in ('scan', 'mixed') else 150
    dpi = st.select_slider(
        "🔍 Resolusi gambar (DPI)",
        options=[100, 120, 150, 200, 250],
        value=default_dpi,
        help="Scan: 200+ disarankan | Digital: 150 cukup",
    )

    # Preprocessing untuk scan
    is_scan_mode = st.session_state.pdf_type in ('scan', 'mixed')
    show_preprocess = st.checkbox(
        "🖼️ Aktifkan image preprocessing",
        value=is_scan_mode,
        help="Tingkatkan kontras & ketajaman gambar sebelum dikirim ke Gemini. Sangat berguna untuk scan.",
    )

    preprocess_mode = 'none'
    if show_preprocess:
        preprocess_mode = st.select_slider(
            "Intensitas preprocessing",
            options=['light', 'balanced', 'strong'],
            value='balanced',
            help="light = ringan | balanced = standar | strong = untuk scan buram",
        )

    # Range halaman
    p_start, p_end = 1, max(st.session_state.n_pages, 1)
    if uploaded and st.session_state.n_pages > 0:
        use_range = st.checkbox("📌 Proses range halaman tertentu")
        if use_range:
            p_start, p_end = st.slider(
                "Range halaman", min_value=1,
                max_value=st.session_state.n_pages,
                value=(1, min(10, st.session_state.n_pages)),
            )

    delay_s = st.slider(
        "⏱️ Jeda antar batch (detik)",
        min_value=1, max_value=10, value=3,
        help="Mencegah rate limit Gemini",
    )

    st.markdown("---")

    btn_proses = st.button(
        "▶️ Mulai Proses",
        type="primary",
        use_container_width=True,
        disabled=not (api_key and uploaded),
    )
    btn_reset = st.button("🗑️ Reset Semua", use_container_width=True)
    if btn_reset:
        st.session_state.entries = []
        st.session_state.feed    = []
        st.session_state.done    = False
        st.rerun()

    st.markdown("---")
    st.markdown(
        "<small style='color:#888'>Font Arab di Excel:<br><b>Noto Sans Arabic</b><br>"
        "<a href='https://fonts.google.com/noto/specimen/Noto+Sans+Arabic' "
        "target='_blank'>Download di sini</a></small>",
        unsafe_allow_html=True,
    )

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN — Stats & Status
# ═══════════════════════════════════════════════════════════════════════════════
stats_ph   = st.empty()
stats_ph.markdown(stats_html_render(st.session_state.entries), unsafe_allow_html=True)

progress_ph = st.empty()
status_ph   = st.empty()

if not st.session_state.done:
    status_ph.markdown(
        '<span class="status-pill status-idle">⬜ Belum diproses</span>',
        unsafe_allow_html=True,
    )
else:
    n = len(st.session_state.entries)
    status_ph.markdown(
        f'<span class="status-pill status-done">✅ Selesai — {n} entri diekstrak</span>',
        unsafe_allow_html=True,
    )

st.markdown("---")

# ── Live feed + Tabel ────────────────────────────────────────────────────────
lcol, rcol = st.columns([1, 2], gap="large")
with lcol:
    st.markdown('<div class="section-title">📡 Live Feed</div>', unsafe_allow_html=True)
    st.caption("Entri baru muncul satu per satu")
    feed_ph = st.empty()

with rcol:
    st.markdown('<div class="section-title">📋 Tabel Semua Entri</div>', unsafe_allow_html=True)
    table_ph = st.empty()

# Render data yang sudah ada
if st.session_state.feed:
    feed_ph.markdown(
        '<div class="feed-wrap">' +
        ''.join(entry_card_html(e) for e in st.session_state.feed[-20:]) +
        '</div>', unsafe_allow_html=True,
    )
if st.session_state.entries:
    table_ph.dataframe(
        pd.DataFrame(st.session_state.entries),
        use_container_width=True, height=420, hide_index=True,
    )

# ── Download ─────────────────────────────────────────────────────────────────
dl_ph = st.empty()
if st.session_state.done and st.session_state.entries:
    with dl_ph.container():
        st.markdown("---")
        _, c2, _ = st.columns([1, 2, 1])
        with c2:
            excel_bytes = make_excel(st.session_state.entries)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.download_button(
                "⬇️ Download Excel (.xlsx)", data=excel_bytes,
                file_name=f"kamus_arab_indonesia_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", use_container_width=True,
            )
        n_l = sum(1 for e in st.session_state.entries if e.get('LEMA'))
        n_s = sum(1 for e in st.session_state.entries if e.get('SUBLEMA'))
        st.success(
            f"✅ **{len(st.session_state.entries)} entri** diekstrak "
            f"({n_l} LEMA + {n_s} SUBLEMA)"
        )

# ═══════════════════════════════════════════════════════════════════════════════
# PROCESSING LOGIC
# ═══════════════════════════════════════════════════════════════════════════════
if btn_proses and api_key and st.session_state.pdf_bytes:

    genai.configure(api_key=api_key)
    model = genai.GenerativeModel(model_name)

    # Apakah mode scan / preprocessing aktif?
    is_scan    = st.session_state.pdf_type in ('scan', 'mixed') or show_preprocess
    prep_mode  = preprocess_mode if show_preprocess else 'none'

    # Range halaman
    if 'use_range' in dir() and use_range:
        idx_start, idx_end = p_start - 1, p_end
    else:
        idx_start, idx_end = 0, st.session_state.n_pages

    batch_starts  = list(range(idx_start, idx_end, batch_size))
    total_batches = len(batch_starts)

    # Reset untuk proses baru
    st.session_state.entries = []
    st.session_state.feed    = []
    st.session_state.done    = False

    status_ph.markdown(
        '<span class="status-pill status-processing">🔄 Sedang diproses...</span>',
        unsafe_allow_html=True,
    )

    for b_idx, b_start in enumerate(batch_starts):
        b_end = min(b_start + batch_size, idx_end)

        pct = b_idx / total_batches
        scan_tag = " [OCR scan]" if is_scan else ""
        progress_ph.progress(
            pct,
            text=f"⏳{scan_tag} Halaman {b_start+1}–{b_end} / {idx_end}  "
                 f"(batch {b_idx+1}/{total_batches})",
        )

        # Konversi halaman → gambar (dengan preprocessing jika scan)
        try:
            imgs = pdf_pages_to_images(
                st.session_state.pdf_bytes,
                b_start, b_end, dpi,
                is_scan=is_scan,
                preprocess_mode=prep_mode,
            )
        except Exception as ex:
            st.warning(f"⚠️ Gagal baca hal. {b_start+1}–{b_end}: {ex}")
            continue

        if not imgs:
            continue

        # Panggil Gemini (mode OCR jika scan)
        new_entries = call_gemini(model, imgs, is_scan=is_scan)

        # Tampilkan satu per satu
        for entry in new_entries:
            st.session_state.entries.append(entry)
            st.session_state.feed.append(entry)

            feed_ph.markdown(
                '<div class="feed-wrap">' +
                ''.join(entry_card_html(e) for e in st.session_state.feed[-20:]) +
                '</div>', unsafe_allow_html=True,
            )
            table_ph.dataframe(
                pd.DataFrame(st.session_state.entries),
                use_container_width=True, height=420, hide_index=True,
            )
            stats_ph.markdown(
                stats_html_render(st.session_state.entries, f"{b_end}/{idx_end}"),
                unsafe_allow_html=True,
            )
            time.sleep(0.06)

        if b_idx < total_batches - 1:
            time.sleep(delay_s)

    progress_ph.progress(1.0, text="✅ Semua batch selesai!")
    st.session_state.done = True
    st.rerun()
